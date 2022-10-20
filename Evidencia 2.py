#IMPORTACIONES
from ast import Try   
import datetime 
from datetime import time, timedelta
import csv 
from collections import namedtuple
import openpyxl
from openpyxl import Workbook
import pandas as pd
from pandas import ExcelWriter
#AGREGAR AL CLIENTE
clientes={}
def agregar_cliente():
    print("***AÑADIR CLIENTE***")
    print('*'*30)
    global clientes
    while True:
        nombre_cliente=input("Ingresa el nombre del cliente:\n")
        if nombre_cliente=="":
            break
        else:
            descripcion=input("Ingrese la empresa a la que pertenece \n")
            if clientes.keys():
                nueva_llave=(max(list(clientes.keys()))+1)
            else:
                nueva_llave=1
            clientes[nueva_llave]=(nombre_cliente,descripcion)
    for id,valores in list(clientes.items()):
        print(f"ID_CLIENTE: {id}\t NOMBRE CLIENTE: {valores[0]:^10}")

#AGREGAR SALA
salas={}
def Agregar_salas():
    print("***REGISTRO DE UNA NUEVA SALA***")
    print('*'*30)
    global salas,datos_sala
    while True:
        nombre_sala=input("Ingresa el nombre de la sala. (Deja en blanco para finalizar)\n")
        if nombre_sala=="":
            break
        else:
            cupo_sala=int(input("Ingresa el cupo de la sala: \n"))
            if cupo_sala == 0:
                print("El cupo no puede ser 0")
            if salas.keys():
                nueva_llave=(max(list(salas.keys()))+1)
            else:
                nueva_llave=1
            salas[nueva_llave]=(nombre_sala,cupo_sala)
    for id_sala,datos_sala in list(salas.items()):
        print(f"ID SALA: {id_sala}\t NOMBRE: {datos_sala[0]:^10}\t CUPO SALA: {datos_sala[1]:^10}")

#REGISTRAR RESERVACION
Reservaciones={}
Turnos={1:'MATUTINO',2:'VESPERTINO',3:'NOCTURNO'}
def Reservar_sala():
    print("RESERVACION DE UNA SALA")
    print('/'*30)
    global Reservaciones
    excepciones=0
    ingresa_id_cliente=int(input("Ingresa el id de cliente \n"))
    if ingresa_id_cliente in clientes:
        registrar_sala=int(input("Ingresa el id de la sala \n"))
        if registrar_sala in salas:
            reservacion_turno=int(input("Ingresa el turno 1 2 3:  "))
            if reservacion_turno>3 or reservacion_turno==0:
                print("Solo se puede escoger 1,2,3")
            if reservacion_turno==1:
                reservacion_turno=Turnos[1]
            if reservacion_turno==2:
                reservacion_turno=Turnos[2]
            if reservacion_turno==3:
                reservacion_turno=Turnos[3]
            if reservacion_turno=="MATUTINO" or reservacion_turno=="VESPERTINO" or reservacion_turno=="NOCTURNO":
                reservacion_evento=input("Ingrese el nombre del evento \n")
                if reservacion_evento=="":
                    print("Error \n")
                else:
                  fecha_evento=input("Ingresa la fecha del evento: \n")
                  fecha_procesada= datetime.datetime.strptime(fecha_evento, "%d/%m/%Y").date()
                  fecha_actual = datetime.date.today()
                  fecha_permitida= fecha_procesada-timedelta(days=2)
                  if fecha_actual<=fecha_permitida:
                    for cliente,sala,turno,fecha,evento,nom in Reservaciones.values():
                      if sala==registrar_sala and turno==reservacion_turno and fecha==fecha_procesada:
                        excepciones=+1     
                    if excepciones==0:
                        nombre_cliente=clientes[ingresa_id_cliente][0]
                        clave= max(list(Reservaciones.keys()),default=0)+1
                        print(f'Su folio es',clave)
                        Reservaciones[clave]=[ingresa_id_cliente,registrar_sala,reservacion_turno,fecha_procesada,reservacion_evento,nombre_cliente]
                    else:
                        print(f'No se puede repetir el mismo turno para una sala')
            else:
                print(f'Solo escoge 1,2 o 3')

#CONSULTAR LAS RESERVACIONES DE UNA SALA EXISTENTE
def consultas():

    fecha_consulta=input("Ingresa la fecha de consulta: \n")
    fecha_procesada01= datetime.datetime.strptime(fecha_consulta, "%d/%m/%Y").date()

    print("\n" + "*"*77)
    print("**" + " "*13 + f"REPORTE DE RESERVACIONES PARA EL DÍA {fecha_consulta}" + " "*13 + "**")
    print("*"*77)
    print("{:<6} {:<20} {:<38} {:<13}".format('SALA','CLIENTE','EVENTO', 'TURNO'))
    print("*"*77)
    for id_reservacion,[cliente_id,sala,turno,fecha,reservacion_evento,nombre_cliente] in Reservaciones.items():
            if fecha_procesada01==fecha:
                print("{:<6} {:<20} {:<38} {:<13}".format(sala, nombre_cliente, reservacion_evento, turno))
    print("*"*30 + " FIN DEL REPORTE " + "*"*30)
                     
#CAMBIAR NOMBRE DEL EVENTO
def cambiar_nombre():
    folio_buscar=int(input("Introduce el folio de la reservacion: "))
    if folio_buscar in Reservaciones:
        Nuevo_nombre=input("Ingresa el nuevo nombre del evento: ")
        Reservaciones[folio_buscar][4]=Nuevo_nombre
        print("Se ha modificado el nombre del evento\t")
 


#CONSULTAR DISPONIBILIDAD DE LAS SALAS
def consultas_d1():
    print("hola")
    lista_encontrados=[]
    reservaciones_realizadas=[]
    if Reservaciones:
        print("Consulta de salas disponibles")
        fecha_consulta=input("Ingresa la fecha de consulta: \n")
        fecha_procesada= datetime.datetime.strptime(fecha_consulta, "%d/%m/%Y").date()
        print("\n" + "*"*77)
        print("**" + " "*13 + f"REPORTE DE SALAS DISPONIBLES PARA EL DÍA {fecha_consulta}" + " "*13 + "**")
        for id_reservacion,[cliente_id,sala,turno,fecha,reservacion_evento,nombre_cliente] in Reservaciones.items():
            if fecha_procesada==fecha:
                for id_sala,datos in list(salas.items()):
                    if sala==id_sala:
                        reservaciones_realizadas.append((id_sala,turno))
        reservaciones_contabilizadas= set(lista_encontrados)
        
        for id_sala,datos in list(salas.items()):
            for clave_turno, nombre in Turnos.items():
                reservaciones_realizadas.append((clave_sala, datos[0], nombre))
        reservas_disponibles = set(reservaciones_realizadas)
              
        turnos_disponibles = sorted(list(reservas_disponibles_ - reservaciones_contabilizadas))

        for clave_sala, sala, turno in turnos_disponibles:
            print("{:<6} {:<10} {:<20}".format(clave_sala, sala, turno))
                        
                    
#EXPORTAR REPORTE TABULAR A EXCEL
def exportar_excel():
    print("***** EXPORTACION A EXCEL *****")
    fecha_consulta=input("Ingresa la fecha de consulta: \n")
    fecha_procesada= datetime.datetime.strptime(fecha_consulta, "%d/%m/%Y").date()
    libro = openpyxl.Workbook()
    libro.iso_dates = True
    hoja = libro["Sheet"] 
    hoja.title = "Reservaciones"
    hoja.cell(row=1,column = 1).value = "sala"
    hoja.cell(row=1,column = 2).value = "nombre cliente"
    hoja.cell(row=1,column = 3).value = "reservacion evento"
    hoja.cell(row=1,column = 4).value = "turno"
    for llave, [cliente_id,sala,turno,fecha,reservacion_evento,nombre_cliente] in Reservaciones.items():
        if fecha==fecha_procesada:
          hoja.cell(row=2,column = 1).value = sala
          hoja.cell(row=2,column = 2).value = nombre_cliente
          hoja.cell(row=2,column = 3).value = reservacion_evento
          hoja.cell(row=2,column = 4).value = turno
    libro.save('Reporte de consulta.xlsx')
    print("listo")
#MENÚ
while True:
    print("******* MENÚ PRINCIPAL *******")
    print("\t[A] RESERVACIONES")
    print("\t[B] REPORTES")
    print("\t[C] REGISTRAR CLIENTE")
    print("\t[D] REGISTRAR SALA")
    print("\t[E] SALIR")

    try:
        opcion=input("\n¿Qué desea hacer? ")
        if opcion=="a":
              print("******* RESERVACIONES *******")
              print("\t[1] NUEVA RESERVACION")
              print("\t[2] MODIFICAR NOMBRE DEL EVENTO DE UNA RESERVACION")
              print("\t[3] CONSULTAR DISPONIBILIDAD DE SALAS")
              opcion_a=int(input("\nElige una opcion: "))
              if opcion_a==1:
                Reservar_sala()
              elif opcion_a==2:
                cambiar_nombre()
              elif opcion_a==3:
                print("Dispo")
                consultas_d1()
        if  opcion=="b":
              print("******* REPORTES *******")
              print("\t[1] MOSTRAR RESERVACIONES PARA UNA FECHA")
              print("\t[2] EXPORTAR EL REPORTE DE LAS RESERVACIONES A EXCEL")
              opcion_b=int(input("\nElige una opcion: "))                
              if opcion_b==1:
                consultas()
              elif opcion_b==2:
                exportar_excel()
        if opcion=="c":
           agregar_cliente()
            
        if opcion=="d":
           Agregar_salas()
        if opcion=="e":
            df = pd.DataFrame(clientes) 
            df.to_csv('clientes.csv',mode='a',sep=';',decimal=',')
            df = pd.DataFrame(salas) 
            df.to_csv('salas.csv',mode='a',sep=';',decimal=',')
            df = pd.DataFrame(Reservaciones) 
            df.to_csv('reservaciones.csv',mode='a',sep=';',decimal=',')
            print("Saliendo......")
    except:
        print("Ingresa una opción válida")

        
